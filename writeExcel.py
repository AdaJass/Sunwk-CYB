#!/usr/bin/python
# -*- coding: utf-8 -*-
import sys,re
import datetime,time
from openpyxl import Workbook #写入excel使用(支持07)
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, colors

def writedb(db):
    filename = "SunWK_CYB.xlsx"
    sheetname = "VCInvDB"
    wb = Workbook()                            #新建一个文件
    wb_writer = ExcelWriter(workbook = wb)     #用来写入文件
    ws = wb.worksheets[0]                      #新建一个sheet
    ws.title = sheetname                       #定义sheet名称
    
    for row_num in range(1,len(db)+1):      #表示遍历行数(最后一个数字循环不到)
        for col_num in range(1,20+1):        #表示遍历列数
            ws.cell(row=row_num+1,column=col_num).value=db[row_num-1][col_num-1]  
            if row_num <= 200:
                cellstr = chr(col_num+96).title()+str(row_num+1) #用ASCII码值读Cell
                xn = ws[cellstr]
                xn = drawcell(xn,colors.BLACK,colors.WHITE)
    for col_num in range(1,20+1):              #补写表头
        ws.cell(row=1,column=col_num).value=formhead(col_num)
    # ws.cell(row=1,column=30).value = lastpage
    # ws.cell(row=1,column=31).value = datalines
    ws = drawheadrow(ws)                       #补画表头
    wb_writer.save(filename = filename)        #保存写入文件
    
def drawcell(xn,fontcolor,cellcolor):          #画单元格基础函数
    btype = 'thin' #medium, thick
    xn.font = Font(name='微软雅黑', size=12, color=fontcolor)
    xn.fill = PatternFill(fill_type='solid', start_color="27408B", end_color="27408B")
    xn.border = Border(left=Side(border_style=btype,color=colors.BLACK),\
                       right=Side(border_style=btype,color=colors.BLACK),\
                       top=Side(border_style=btype,color=colors.BLACK),\
                       bottom=Side(border_style=btype,color=colors.BLACK),\
                       outline=Side(border_style=btype,color=colors.BLACK))
    xn.fill = PatternFill(fill_type='solid', start_color=cellcolor)
    return xn

def drawheadrow(ws_n):                         #把表头画得好看些
    a1 = ws_n['A1']
    a1 = drawcell(a1,colors.WHITE,"7A378B")
    b1 = ws_n['B1']
    b1 = drawcell(b1,colors.WHITE,"27408B")
    c1 = ws_n['C1']
    c1 = drawcell(c1,colors.WHITE,"27408B")
    d1 = ws_n['D1']
    d1 = drawcell(d1,colors.WHITE,"8B0000")
    e1 = ws_n['E1']
    e1 = drawcell(e1,colors.WHITE,"EEB422")
    f1 = ws_n['F1']
    f1 = drawcell(f1,colors.WHITE,"EEB422")
    g1 = ws_n['G1']
    g1 = drawcell(g1,colors.WHITE,"218868")
    h1 = ws_n['H1']
    h1 = drawcell(h1,colors.WHITE,"218868")
    i1 = ws_n['I1']
    i1 = drawcell(i1,colors.WHITE,"218868")
    j1 = ws_n['J1']
    j1 = drawcell(j1,colors.WHITE,"218868")
    k1 = ws_n['K1']
    k1 = drawcell(k1,colors.WHITE,"218868")
    l1 = ws_n['L1']
    l1 = drawcell(l1,colors.WHITE,"218868")
    m1 = ws_n['M1']
    m1 = drawcell(m1,colors.WHITE,"AAAAAA")
    n1 = ws_n['N1']
    n1 = drawcell(n1,colors.WHITE,"AAAAAA")
    o1 = ws_n['O1']
    o1 = drawcell(o1,colors.WHITE,"AAAAAA")
    p1 = ws_n['P1']
    p1 = drawcell(p1,colors.WHITE,"AAAAAA")
    q1 = ws_n['Q1']
    q1 = drawcell(q1,colors.WHITE,"AAAAAA")
    r1 = ws_n['R1']
    r1 = drawcell(r1,colors.WHITE,"AAAAAA")
    s1 = ws_n['S1']
    s1 = drawcell(s1,colors.WHITE,"AAAAAA")
    t1 = ws_n['T1']
    t1 = drawcell(t1,colors.WHITE,"AAAAAA")
    return ws_n

def formhead(formindex):
    if formindex == 1:
        formtitle = "融资时间"
    elif formindex == 2:
        formtitle = "项目名称"
    elif formindex == 3:
        formtitle = "所属行业"
    elif formindex == 4:
        formtitle = "公司全名"
    elif formindex == 5:
        formtitle = "融资轮次"
    elif formindex == 6:
        formtitle = "投资金额/万RMB"
    elif formindex == 7:
        formtitle = "投资方1"
    elif formindex == 8:
        formtitle = "投资方2"
    elif formindex == 9:
        formtitle = "投资方3"
    elif formindex == 10:
        formtitle = "投资方4"
    elif formindex == 11:
        formtitle = "投资方5"
    elif formindex == 12:
        formtitle = "投资方6"
    else:
        formtitle = "其他投资方"
    return formtitle


def writeCsv(db):    
    with open('sunwk-cyb.csv', 'w', encoding='utf-8') as f:        
        for col_num in range(1,20+1):              #补写表头            
            f.write(formhead(col_num))
            if col_num != 20:
                f.write(',')
        f.write('\n')
        if len(db)>1000:
            csv_rows=1000
        else:
            csv_rows=len(db)

        for row_num in range(1,csv_rows+1):      #表示遍历行数(最后一个数字循环不到)            
            for col_num in range(1,20+1):        #表示遍历列数
                if not db[row_num-1][col_num-1]:
                    if col_num != 20:
                        f.write(',')                  
                else:
                    f.write(str(db[row_num-1][col_num-1]))
                    if col_num != 20:
                        f.write(',')
            f.write('\n')


if __name__ == '__main__':
    data=[]
    cell=[1,2,3,4,0,0,0,'seikjkgd',0,0,0,0,0,0,0,0,0,0,0,0]
    data.append(cell)
    data.append(cell)
    writedb(data)
    writeCsv(data)
